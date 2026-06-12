import csv
import json
import re
import os

def normalize_text(text):
    if not text:
        return ''
    text = text.upper()
    replacements = {'Á':'A', 'É':'E', 'Í':'I', 'Ó':'O', 'Ú':'U', 'Ñ':'N', 'Ü':'U'}
    for k, v in replacements.items():
        text = text.replace(k, v)
    # Remove common abbreviations and indicators of education center
    text = re.sub(r'\b(I\.?E\.?R?|C\.?E\.?R?|S\.?D|I\.?E\.?S\.?E\.?D?|E\.?U|I\.?E|COL\.?|COLLEGIO|COLEGIO|ESCUELA|CENTRO EDUCATIVO|INSTITUCION EDUCATIVA|SEDE)\b', '', text)
    # Keep only letters and numbers
    text = re.sub(r'[^A-Z0-9]', '', text)
    return text

def get_words(name):
    n = name.upper()
    # Replace synonyms
    n = re.sub(r'\b(P\.?\s*BIBL\.?|PARQUE\s+BIBLIOTECA|PAR\.?\s*BIBLIOTECA)\b', 'BIBLIOTECA', n)
    n = re.sub(r'\b(SEC\.?\s*ESC\.?|SECCION\s+ESCUELA|ESCUELA)\b', 'ESCUELA', n)
    n = re.sub(r'\b(I\.?E\.?|INSTITUCION\s+EDUCATIVA)\b', 'IE', n)
    n = re.sub(r'\b(COL\.?|COLEGIO)\b', 'COLEGIO', n)
    n = re.sub(r'\b(SEC\.?\s*COL\.?|SECCION\s+COLEGIO)\b', 'COLEGIO', n)
    n = re.sub(r'\b(INST\.?|INSTITUTO)\b', 'INSTITUTO', n)
    n = re.sub(r'\b(SEC\.?\s*ESC\.?\s*AUX\.?)\b', 'ESCUELA', n)
    n = re.sub(r'\b(CORP\.?|CORPORACION)\b', 'CORPORACION', n)
    n = re.sub(r'\b(UNIV\.?|UNIVERSIDAD)\b', 'UNIVERSIDAD', n)
    # Remove accents
    replacements = {'Á':'A', 'É':'E', 'Í':'I', 'Ó':'O', 'Ú':'U', 'Ñ':'N', 'Ü':'U'}
    for k, v in replacements.items():
        n = n.replace(k, v)
    # Extract words of length >= 3
    words = re.findall(r'[A-Z0-9]{3,}', n)
    return set(words)

manual_overrides = {
    # Medellín posts
    "I.E. SANTA JUANA DE LESTONNAC": (6.30304, -75.5755),
    "I.E. JORGE ELIECER GAITAN": (6.28991, -75.5852),
    "I.E. VILLA DE LA CANDELARIA": (6.285044, -75.593645),
    "INSTITUTO FERRINI - SEDE ROBLEDO": (6.27792, -75.5923),
    "I.T.M. (CAMPUS FRATERNIDAD)": (6.2453704, -75.5510349),
    "I.E. MANUEL JOSÉ CAYZEDO": (6.24144, -75.55199),
    "I.E. MANUEL JOSE CAYZEDO": (6.24144, -75.55199),
    "AUDITORIO CC SAN DIEGO TORRE NTE PISO 11": (6.2341186, -75.5686026),
    "PLAZA MINORISTA JOSE MARIA VILLA": (6.2576437, -75.5734711),
    "I.E. MARCO FIDEL SUAREZ": (6.2520, -75.5860),
    "COL. CALASANZ (FEM) CAMPESTRE ESCOLAPIAS": (6.2665, -75.5903),
    "COLEGIO BETHLEMITAS": (6.2408545, -75.5955852),
    "CORP. UNIVERSITARIA ADVENTISTA - UNAC": (6.24107, -75.6094),
    "COLEGIO LATINO": (6.2430826, -75.6113314),
    "C4TA CIUDADELA": (6.2526, -75.6124),
    "SEC. ESC. BETANIA": (6.24622, -75.56227),
    "I.E. CARLOS VIECO ORTIZ": (6.25479, -75.6199),
    "SEC. ESC. AMOR AL NIÑO": (6.25304, -75.6245),
    "SEC. ESC. AMOR AL NIO": (6.25304, -75.6245),
    "I.E FUNDADORES": (6.2575, -75.6145),
    "SEC. ESC. VEINTE DE JULIO": (6.25229, -75.62084),
    "PARQUE COMERCIAL EL TESORO": (6.1974, -75.5586),
    "I.E. LA SALLE DE CAMPOAMOR": (6.21267, -75.5876),
    "SEC. ESC. SANTISIMA TRINIDAD": (6.2185, -75.5802),
    "INSTITUTO SAN CARLOS": (6.22315, -75.6001),
    "I.E. CAPILLA DEL ROSARIO": (6.20518, -75.6021),
    "C.E. MARIA PAULINA TABORDA": (6.2415, -75.6425),
    "COLEGIO PACHA MAMA": (6.2307, -75.6288),
    "I.E. MANUEL J. BETANCUR": (6.183666773, -75.65749518),
    "C.E. JUAN ANDRES PATIÑO": (6.2319789, -75.4799755),
    "C.E. JUAN ANDRES PATIO": (6.2319789, -75.4799755),
    "C.E. MEDIA LUNA": (6.2305, -75.5135),
    "I.E. PBRO ANTONIO JOSE BERNAL LONDOÑO SJ": (6.301311, -75.559322),
    "I.E. PBRO ANTONIO JOSE BERNAL LONDONO SJ": (6.301311, -75.559322),
    "I.E. MONSEÑOR FRANCISCO CRISTOBAL TORO": (6.28255, -75.560399),
    "I.E. MONSENOR FRANCISCO CRISTOBAL TORO": (6.28255, -75.560399),
    "I.E. JOSE EUSEBIO CARO": (6.285550000000001, -75.558786929),
    "I.E. NUEVO HORIZONTE - PAULO VI": (6.2947, -75.5460),
    "SEC. ESC. NUEVO HORIZONTE 2": (6.2947, -75.5460),
    "SEC. ESC. BALDOMERO SANIN CANO": (6.2731, -75.5524),
    "IE RODRIGO LARA BONILLA - SEDE LA SUSANA": (6.26243, -75.553199),
}


def parse_coord(coord_str, is_lon=False):
    if not coord_str:
        return None
    s = coord_str.strip().replace(' ', '')
    if 'E' in s or 'e' in s:
        s_dot = s.replace(',', '.')
        try:
            val = float(s_dot)
            abs_val = abs(val)
            if is_lon:
                while abs_val >= 78.0:
                    abs_val /= 10.0
                while abs_val < 73.0:
                    abs_val *= 10.0
                return -abs_val
            else:
                while abs_val >= 9.0:
                    abs_val /= 10.0
                while abs_val < 4.0:
                    abs_val *= 10.0
                return abs_val
        except ValueError:
            pass
    
    clean = re.sub(r'[^0-9\-]', '', s)
    if not clean:
        return None
    try:
        clean_str = clean.lstrip('-')
        if is_lon:
            # We want something around -74 to -77
            if len(clean_str) >= 2:
                # E.g. -75541166 -> -75.541166
                return -float(clean_str[0:2] + '.' + clean_str[2:])
        else:
            # We want something around 5 to 8
            if len(clean_str) >= 1:
                return float(clean_str[0:1] + '.' + clean_str[1:])
    except Exception:
        pass
        
    try:
        val = float(s.replace(',', '.'))
        if is_lon:
            return val if val < 0 else -val
        return val
    except ValueError:
        return None

# Candidates list (order must match original)
candidates = [
    "ABELARDO DE LA ESPRIELLA",
    "CLAUDIA LÓPEZ",
    "GUSTAVO MATAMOROS CAMACHO",
    "IVÁN CEPEDA CASTRO",
    "MIGUEL URIBE LONDOÑO",
    "PALOMA VALENCIA LASERNA",
    "RAÚL SANTIAGO BOTERO JARAMILLO",
    "ROY LEONARDO BARRERAS MONTEALEGRE",
    "SERGIO FAJARDO VALDERRAMA",
    "SONDRA MACOLLINS GARVIN PINTO",
    "VOTOS EN BLANCO",
    "VOTOS NO MARCADOS",
    "VOTOS NULOS",
    "ÓSCAR MAURICIO LIZCANO ARANGO"
]

candidate_mapping = {
    "ABELARDO DE LA ESPRIELLA": "ABELARDO DE LA ESPRIELLA",
    "CLAUDIA LOPEZ": "CLAUDIA LÓPEZ",
    "CLAUDIA LÓPEZ": "CLAUDIA LÓPEZ",
    "GUSTAVO MATAMOROS CAMACHO": "GUSTAVO MATAMOROS CAMACHO",
    "IVAN CEPEDA CASTRO": "IVÁN CEPEDA CASTRO",
    "IVÁN CEPEDA CASTRO": "IVÁN CEPEDA CASTRO",
    "MIGUEL URIBE LONDONO": "MIGUEL URIBE LONDOÑO",
    "MIGUEL URIBE LONDOÑO": "MIGUEL URIBE LONDOÑO",
    "PALOMA VALENCIA LASERNA": "PALOMA VALENCIA LASERNA",
    "RAUL SANTIAGO BOTERO JARAMILLO": "RAÚL SANTIAGO BOTERO JARAMILLO",
    "RAÚL SANTIAGO BOTERO JARAMILLO": "RAÚL SANTIAGO BOTERO JARAMILLO",
    "ROY LEONARDO BARRERAS MONTEALEGRE": "ROY LEONARDO BARRERAS MONTEALEGRE",
    "SERGIO FAJARDO VALDERRAMA": "SERGIO FAJARDO VALDERRAMA",
    "SONDRA MACOLLINS GARVIN PINTO": "SONDRA MACOLLINS GARVIN PINTO",
    "VOTOS EN BLANCO": "VOTOS EN BLANCO",
    "VOTOS NO MARCADOS": "VOTOS NO MARCADOS",
    "VOTOS NULOS": "VOTOS NULOS",
    "OSCAR MAURICIO LIZCANO ARANGO": "ÓSCAR MAURICIO LIZCANO ARANGO",
    "ÓSCAR MAURICIO LIZCANO ARANGO": "ÓSCAR MAURICIO LIZCANO ARANGO"
}

def clean_cand_name(name):
    # Normalize candidate name from CSV to map to standard list
    n = name.upper().strip()
    # Remove accents
    replacements = {'Á':'A', 'É':'E', 'Í':'I', 'Ó':'O', 'Ú':'U', 'Ñ':'N', 'Ü':'U'}
    for k, v in replacements.items():
        n = n.replace(k, v)
    # Map to final name
    if n in candidate_mapping:
        return candidate_mapping[n]
    # Try fuzzy mapping or prefix matching
    for key, val in candidate_mapping.items():
        if key in n or n in key:
            return val
    return n

# 1. Load coordinates from divipole and Coordenadas files
divipole_coords = {} # (muni_norm, post_norm) -> (lat, lon)
divipole_coords_raw = [] # list of (muni_norm, puesto_raw, lat, lon)
muni_coords = {} # muni_norm -> list of (lat, lon)

def load_coords_from_file(filepath):
    print(f"Loading coordinates from {filepath}...")
    with open(filepath, mode='r', encoding='latin1') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            muni = row.get('MUNICIPIO', '').strip()
            puesto = row.get('PUESTO', '').strip()
            direccion = row.get('DIRECCIÓN', row.get('DIRECCIN', '')).strip()
            departamento = row.get('DEPARTAMENTO', '').strip().upper()
            
            # Skip if there's a department specified and it's not Antioquia
            if departamento and departamento != 'ANTIOQUIA':
                continue
                
            lat_raw = row.get('LATITUD', '')
            lon_raw = row.get('LONGITUD', '')
            
            lat = parse_coord(lat_raw, False)
            lon = parse_coord(lon_raw, True)
            
            if lat is not None and lon is not None:
                muni_norm = normalize_text(muni)
                puesto_norm = normalize_text(puesto)
                
                # Store raw for fuzzy matching
                divipole_coords_raw.append((muni_norm, puesto, lat, lon))
                
                # Store
                key = (muni_norm, puesto_norm)
                if key not in divipole_coords:
                    divipole_coords[key] = (lat, lon)
                
                # Also store by address
                dir_norm = normalize_text(direccion)
                if dir_norm:
                    divipole_coords[(muni_norm, dir_norm)] = (lat, lon)
                
                if muni_norm not in muni_coords:
                    muni_coords[muni_norm] = []
                muni_coords[muni_norm].append((lat, lon))


load_coords_from_file('divipole_puestos_antioquia.csv')
if os.path.exists('Coordenadas Puestos Votacion.csv'):
    load_coords_from_file('Coordenadas Puestos Votacion.csv')

# Calculate average coordinates for each municipality
muni_avg_coords = {}
for muni_norm, coords_list in muni_coords.items():
    lats = [c[0] for c in coords_list]
    lons = [c[1] for c in coords_list]
    muni_avg_coords[muni_norm] = (sum(lats)/len(lats), sum(lons)/len(lons))

# 2. Process MMV file and aggregate data
print("Processing MMV results...")

hierarchy = {}
communes_votes = {} # commune_id -> candidate -> votes

# Medellin commune mapping
commune_mapping = {
    "001": "01",
    "002": "02",
    "003": "03",
    "004": "04",
    "005": "05",
    "006": "06",
    "007": "07",
    "008": "08",
    "009": "09",
    "010": "10",
    "011": "11",
    "012": "12",
    "013": "13",
    "014": "14",
    "015": "15",
    "016": "16",
    "017": "70", # Altavista
    "018": "80", # San Antonio de Prado
    "019": "50", # Palmitas
    "020": "60", # San Cristobal
    "021": "90"  # Santa Elena
}

with open('MMV_XXX_01_000_XXX_XX_XX_XXX_1000.csv', mode='r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter=';')
    for row in reader:
        muni = row['MUNNOMBRE'].strip().upper()
        muni_norm = normalize_text(muni)
        zone = f"Zona {int(row['ZONA'])}"
        post_id = row['PUESTO'].zfill(2)
        post_name = row['PUESNOMBRE'].strip()
        table = str(int(row['MESA']))
        candidate_raw = row['CANNOMBRE'].strip()
        candidate = clean_cand_name(candidate_raw)
        votes = int(row['VOTOS'])
        
        # Check candidate validity
        if candidate not in candidates:
            # Try to log or ignore
            continue
            
        cand_idx = candidates.index(candidate)
        
        # Hierarchy aggregation
        if muni not in hierarchy:
            hierarchy[muni] = {}
        if zone not in hierarchy[muni]:
            hierarchy[muni][zone] = {}
        if post_id not in hierarchy[muni][zone]:
            # Try to match coordinate
            lat, lon = None, None
            post_norm = normalize_text(post_name)
            
            # Match heuristics
            # 1. Manual overrides first
            if post_name in manual_overrides:
                lat, lon = manual_overrides[post_name]
            # 2. Exact match in database
            elif (muni_norm, post_norm) in divipole_coords:
                lat, lon = divipole_coords[(muni_norm, post_norm)]
            else:
                # 3. Fuzzy match using word overlap ratio
                post_words = get_words(post_name)
                best_match = None
                best_overlap = 0.0
                
                for m_n, p_raw, la, lo in divipole_coords_raw:
                    if m_n == muni_norm:
                        p_words = get_words(p_raw)
                        intersection = post_words.intersection(p_words)
                        if intersection:
                            ratio = len(intersection) / min(len(post_words), len(p_words))
                            if ratio > best_overlap:
                                best_overlap = ratio
                                best_match = (la, lo)
                
                if best_match and best_overlap >= 0.75:
                    lat, lon = best_match
                else:
                    # 4. Substring matching
                    matched_coord = None
                    for (m_n, p_n), (la, lo) in divipole_coords.items():
                        if m_n == muni_norm:
                            if p_n in post_norm or post_norm in p_n:
                                matched_coord = (la, lo)
                                break
                    if matched_coord:
                        lat, lon = matched_coord
                    else:
                        # Unmatched: DO NOT fall back to municipal average to avoid false spatial joins
                        lat, lon = None, None
            
            if lat is None or lon is None:
                print(f"WARNING: Unmatched post '{post_name}' in {muni}")

            
            hierarchy[muni][zone][post_id] = {
                "name": post_name,
                "lat": lat,
                "lon": lon,
                "tables": {}
            }
            
        post_obj = hierarchy[muni][zone][post_id]
        if table not in post_obj["tables"]:
            post_obj["tables"][table] = [0] * len(candidates)
            
        post_obj["tables"][table][cand_idx] += votes
        
        # Commune aggregation (only for Medellin and valid commune codes)
        if muni == 'MEDELLIN':
            comm_code = row['COMUCODIGO']
            if comm_code in commune_mapping:
                comm_id = commune_mapping[comm_code]
                if comm_id not in communes_votes:
                    communes_votes[comm_id] = {c: 0 for c in candidates}
                communes_votes[comm_id][candidate] += votes

# 3. Format Medellín Communes object
communes_output = {}
for comm_id, votes_dict in communes_votes.items():
    total_votes = sum(votes_dict.values())
    results_list = []
    for cand, v in votes_dict.items():
        results_list.append({
            "candidate": cand,
            "votes": v,
            "pct": (v / total_votes * 100) if total_votes > 0 else 0
        })
    results_list.sort(key=lambda x: x["votes"], reverse=True)
    
    winner = results_list[0]["candidate"]
    winner_votes = results_list[0]["votes"]
    winner_pct = results_list[0]["pct"]
    
    communes_output[comm_id] = {
        "total_votes": total_votes,
        "winner": winner,
        "winner_votes": winner_votes,
        "winner_pct": winner_pct,
        "results": results_list
    }

output_data = {
    "candidates": candidates,
    "communes": communes_output,
    "hierarchy": hierarchy
}

# 4. Load 2022 data for comparison (Dynamic)
muni_2022 = {}
hierarchy_2022 = {}

try:
    with open('MMV_ANTIOQUIA_2022_1v.csv', mode='r', encoding='latin1') as f:
        reader = csv.DictReader(f, delimiter=',')
        for row in reader:
            muni = row.get('MUNNOMBRE', '').strip().upper()
            muni_norm = normalize_text(muni)
            zone_num = row.get('ZONA', '0').strip()
            zone = f"Zona {int(zone_num)}" if zone_num.isdigit() else "Zona 0"
            post_id = row.get('PUESTO', '0').strip().zfill(2)
            
            cand = row.get('CANNOMBRE', '').strip().upper()
            replacements = {'Á':'A', 'É':'E', 'Í':'I', 'Ó':'O', 'Ú':'U', 'Ñ':'N', 'Ü':'U'}
            for k, v in replacements.items():
                cand = cand.replace(k, v)
                
            votes = int(row.get('VOTOS', 0))
            
            # Filter to keep JSON small: only Fico, Petro, and Total
            # We will use "FICO", "PETRO", and "TOTAL" as keys
            key = None
            if 'FEDERICO' in cand or 'RODOLFO HERN' in cand:
                key = 'FICO'
            elif 'GUSTAVO PETRO' in cand:
                key = 'PETRO'
            else:
                key = 'OTHER'
                
            # Muni totals
            if muni not in muni_2022:
                muni_2022[muni] = {'FICO': 0, 'PETRO': 0, 'TOTAL': 0}
            if key in ['FICO', 'PETRO']:
                muni_2022[muni][key] += votes
            muni_2022[muni]['TOTAL'] += votes
            
            # Hierarchy totals
            if muni not in hierarchy_2022:
                hierarchy_2022[muni] = {}
            if zone not in hierarchy_2022[muni]:
                hierarchy_2022[muni][zone] = {}
            if post_id not in hierarchy_2022[muni][zone]:
                hierarchy_2022[muni][zone][post_id] = {'FICO': 0, 'PETRO': 0, 'TOTAL': 0}
            
            if key in ['FICO', 'PETRO']:
                hierarchy_2022[muni][zone][post_id][key] += votes
            hierarchy_2022[muni][zone][post_id]['TOTAL'] += votes

    # Store aggregated 2022 data
    output_data["comparison_2022_muni"] = muni_2022
    output_data["comparison_2022_hierarchy"] = hierarchy_2022
    
except Exception as e:
    print(f"Could not load 2022 data for comparison: {e}")


# 4. Load potentials from Potencial.xlsx
print("Loading potentials from Potencial.xlsx...")
zone_potentials = {}
try:
    import openpyxl
    wb = openpyxl.load_workbook('Potencial.xlsx')
    ws = wb.active
    for name in wb.sheetnames:
        if 'Divisi' in name or 'Electoral' in name:
            ws = wb[name]
            break
    for row in ws.iter_rows(min_row=2):
        desc = row[0].value
        pot = row[1].value
        if desc and pot and "TOTAL" not in str(desc).upper():
            zone_num = str(desc).replace("Zona ", "")
            try:
                zone_key = f"Zona {int(zone_num)}"
                zone_potentials[zone_key] = int(pot)
            except ValueError:
                pass
    output_data["zone_potentials"] = zone_potentials
    print(f"Successfully loaded potentials for {len(zone_potentials)} zones.")
except Exception as e:
    print(f"Error loading Potencial.xlsx: {e}")

# 4.5 Load exact potentials per post from Divipole Antioquia.xlsx
print("Loading exact post potentials from Divipole Antioquia.xlsx...")
post_potentials = {}
try:
    import pandas as pd
    df_divipole = pd.read_excel('Divipole Antioquia.xlsx')
    for _, row in df_divipole.iterrows():
        puesto_raw = str(row.get('PUESTO', ''))
        potencial = row.get('TOTAL', 0)
        puesto_norm = normalize_text(puesto_raw)
        
        # Guardar bajo diferentes formas para facilitar el match luego
        post_potentials[puesto_norm] = int(potencial)
        post_potentials[puesto_raw] = int(potencial)
        
    output_data["post_potentials"] = post_potentials
    print(f"Successfully loaded exact potentials for {len(post_potentials)} posts.")
except Exception as e:
    print(f"Error loading Divipole Antioquia.xlsx: {e}")

# 5. Save to electoral_data.json
print("Saving output to electoral_data.json...")
with open('electoral_data.json', 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print("Data processing complete!")

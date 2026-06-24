import openpyxl
import os
import unicodedata

def normalize(texto):
    if not texto: return "Sin_Nombre"
    texto = str(texto).strip()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = texto.replace('', '') # Remove weird characters if any
    return texto.strip()

print("Cargando el archivo Excel... Esto puede tomar un momento.")
wb = openpyxl.load_workbook('REPRESENTANTES 2026-2030 CON FOTO UV.xlsx')
ws = wb.active

dest_dir = 'fotos_extraidas_representantes'
os.makedirs(dest_dir, exist_ok=True)

images_extracted = 0

for idx, img in enumerate(ws._images):
    # img.anchor._from.row es 0-indexed. 
    # Openpyxl cells son 1-indexed (row=1, col=1)
    row_idx = img.anchor._from.row + 1
    
    # Columna C (3) es APELLIDOS, Columna D (4) es NOMBRES
    apellidos = ws.cell(row=row_idx, column=3).value
    nombres = ws.cell(row=row_idx, column=4).value
    
    # Algunas veces las celdas pueden estar desplazadas o ser None
    # Si ambas son None, probemos la siguiente fila por si acaso la imagen está anclada a la celda superior
    if not apellidos and not nombres:
        apellidos = ws.cell(row=row_idx+1, column=3).value
        nombres = ws.cell(row=row_idx+1, column=4).value
        
    nombres_str = normalize(nombres) if nombres else ""
    apellidos_str = normalize(apellidos) if apellidos else ""
    
    full_name = f"{nombres_str} {apellidos_str}".strip()
    if not full_name:
        full_name = f"Representante_Desconocido_{idx}"
        
    # Limpiar nombre para archivo
    safe_name = "".join([c for c in full_name if c.isalpha() or c.isdigit() or c==' ']).rstrip()
    
    ext = 'jpeg' # default
    if hasattr(img, 'format'):
        ext = img.format
    elif hasattr(img.ref, 'getvalue'):
        # Pillow image, determine format
        pass # just use jpeg
        
    filename = f"{safe_name}.{ext}"
    filepath = os.path.join(dest_dir, filename)
    
    # Manejar imágenes que son directamente datos binarios
    # En openpyxl, el img._data es una función lambda o bytes directos
    img_data = img._data() if callable(img._data) else img._data
    
    with open(filepath, 'wb') as f:
        f.write(img_data)
        
    images_extracted += 1

print(f"Completado. {images_extracted} fotos extraidas en la carpeta '{dest_dir}'.")
